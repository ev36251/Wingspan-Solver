"""Parse the Wingspan Excel spreadsheet into Python data objects."""

import re
from pathlib import Path

import openpyxl

from backend.config import (
    BONUS_CARD_COLUMNS,
    INCLUDED_SET_NAMES,
)
from backend.models.enums import (
    BeakDirection,
    FoodType,
    GameSet,
    Habitat,
    NestType,
    PowerColor,
    SET_NAME_MAP,
)
from backend.models.bird import Bird, FoodCost
from backend.models.bonus_card import BonusCard, BonusScoringTier
from backend.models.goal import Goal


# Column indices for the Birds sheet (0-indexed)
class BirdCol:
    NAME = 0
    SCIENTIFIC = 1
    SET = 2
    COLOR = 3
    POWER_TEXT = 4
    PREDATOR = 5
    FLOCKING = 6
    BONUS_CARD = 7
    VP = 8
    NEST = 9
    EGG_LIMIT = 10
    WINGSPAN = 11
    FOREST = 12
    GRASSLAND = 13
    WETLAND = 14
    INVERTEBRATE = 15
    SEED = 16
    FISH = 17
    FRUIT = 18
    RODENT = 19
    NECTAR = 20
    WILD = 21
    SLASH = 22  # "/" means OR cost
    STAR = 23   # "*" means predator cost
    TOTAL_COST = 24
    BEAK = 25


def _parse_food_cost(row: tuple) -> FoodCost:
    """Parse food cost from a bird row."""
    items = []

    food_cols = [
        (BirdCol.INVERTEBRATE, FoodType.INVERTEBRATE),
        (BirdCol.SEED, FoodType.SEED),
        (BirdCol.FISH, FoodType.FISH),
        (BirdCol.FRUIT, FoodType.FRUIT),
        (BirdCol.RODENT, FoodType.RODENT),
        (BirdCol.NECTAR, FoodType.NECTAR),
        (BirdCol.WILD, FoodType.WILD),
    ]

    for col_idx, food_type in food_cols:
        count = row[col_idx]
        if count and count not in (None, "", 0, "0"):
            for _ in range(int(float(count))):
                items.append(food_type)

    is_or = bool(row[BirdCol.SLASH] and str(row[BirdCol.SLASH]).strip() == "/")
    total = int(float(row[BirdCol.TOTAL_COST])) if row[BirdCol.TOTAL_COST] else 0

    return FoodCost(items=tuple(items), is_or=is_or, total=total)


def _parse_habitats(row: tuple) -> frozenset[Habitat]:
    """Parse which habitats a bird can live in."""
    habitats = set()
    if row[BirdCol.FOREST] and str(row[BirdCol.FOREST]).strip():
        habitats.add(Habitat.FOREST)
    if row[BirdCol.GRASSLAND] and str(row[BirdCol.GRASSLAND]).strip():
        habitats.add(Habitat.GRASSLAND)
    if row[BirdCol.WETLAND] and str(row[BirdCol.WETLAND]).strip():
        habitats.add(Habitat.WETLAND)
    return frozenset(habitats)


def _parse_nest_type(value) -> NestType:
    """Parse nest type from spreadsheet value."""
    if not value:
        return NestType.WILD
    v = str(value).strip().lower()
    return {
        "bowl": NestType.BOWL,
        "cavity": NestType.CAVITY,
        "ground": NestType.GROUND,
        "platform": NestType.PLATFORM,
        "wild": NestType.WILD,
        "star": NestType.WILD,
        "*": NestType.WILD,
    }.get(v, NestType.WILD)


def _parse_beak(value) -> BeakDirection:
    """Parse beak direction."""
    if not value:
        return BeakDirection.NONE
    v = str(value).strip().upper()
    if v in ("L", "LEFT"):
        return BeakDirection.LEFT
    if v in ("R", "RIGHT"):
        return BeakDirection.RIGHT
    return BeakDirection.NONE


def _parse_color(value) -> PowerColor:
    """Parse power color."""
    if not value:
        return PowerColor.NONE
    v = str(value).strip().lower()
    return {
        "white": PowerColor.WHITE,
        "brown": PowerColor.BROWN,
        "pink": PowerColor.PINK,
        "teal": PowerColor.TEAL,
        "yellow": PowerColor.YELLOW,
    }.get(v, PowerColor.NONE)


def _parse_bonus_eligibility(row: tuple) -> frozenset[str]:
    """Parse which bonus cards a bird qualifies for (columns 39-64)."""
    eligible = set()
    for col_idx, bonus_name in BONUS_CARD_COLUMNS.items():
        val = row[col_idx]
        if val and str(val).strip().upper() == "X":
            eligible.add(bonus_name)
    return frozenset(eligible)


def _parse_wingspan(value) -> int | None:
    """Parse wingspan, returning None for flightless ('*') birds."""
    if not value or str(value).strip() == "*":
        return None
    return int(float(value))


def load_birds(filepath: Path) -> list[Bird]:
    """Load all birds from the Excel file, filtered to included sets."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["Birds"]

    birds = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # Skip header row and two stat rows
        if i < 3:
            continue

        game_set_str = row[BirdCol.SET]
        if not game_set_str or game_set_str not in INCLUDED_SET_NAMES:
            continue

        name = row[BirdCol.NAME]
        if not name:
            continue

        game_set = SET_NAME_MAP[game_set_str]

        bird = Bird(
            name=str(name).strip(),
            scientific_name=str(row[BirdCol.SCIENTIFIC] or "").strip(),
            game_set=game_set,
            color=_parse_color(row[BirdCol.COLOR]),
            power_text=str(row[BirdCol.POWER_TEXT] or "").strip(),
            victory_points=int(float(row[BirdCol.VP] or 0)),
            nest_type=_parse_nest_type(row[BirdCol.NEST]),
            egg_limit=int(float(row[BirdCol.EGG_LIMIT] or 0)),
            wingspan_cm=_parse_wingspan(row[BirdCol.WINGSPAN]),
            habitats=_parse_habitats(row),
            food_cost=_parse_food_cost(row),
            beak_direction=_parse_beak(row[BirdCol.BEAK]),
            is_predator=bool(row[BirdCol.PREDATOR] and str(row[BirdCol.PREDATOR]).strip()),
            is_flocking=bool(row[BirdCol.FLOCKING] and str(row[BirdCol.FLOCKING]).strip()),
            is_bonus_card_bird=bool(row[BirdCol.BONUS_CARD] and str(row[BirdCol.BONUS_CARD]).strip()),
            bonus_eligibility=_parse_bonus_eligibility(row),
        )
        birds.append(bird)

    wb.close()
    return birds


def _parse_bonus_vp(vp_text: str) -> tuple[tuple[BonusScoringTier, ...], bool]:
    """Parse bonus card VP text into scoring tiers.

    Returns (tiers, is_per_bird).

    Patterns handled:
    - "2 per bird" / "1 each" / "3 each" -> per-bird scoring
    - "N per column" / "N per power color" -> per-unit scoring
    - "X to Y birds: P1; Z+ birds: P2" -> tiered
    - "3 consecutive birds: P1; 4 consecutive: P2; 5 consecutive: P3" -> tiered
    - "One set: P1; Two sets: P2" -> tiered
    - "0 to 2 points: 6; 3 points: 3; 4+ points: 0" -> inverted tiered
    """
    if not vp_text or vp_text in ("-", "None", ""):
        return ((), False)

    text = vp_text.strip()

    # "N per bird" or "N each"
    m = re.match(r"(\d+)\s+per\s+bird", text)
    if m:
        pts = int(m.group(1))
        return ((BonusScoringTier(min_count=1, max_count=None, points=pts),), True)

    m = re.match(r"(\d+)\s+each", text)
    if m:
        pts = int(m.group(1))
        return ((BonusScoringTier(min_count=1, max_count=None, points=pts),), True)

    # "N per column" or "N per power color"
    m = re.match(r"(\d+)\s+per\s+(\w+)", text)
    if m:
        pts = int(m.group(1))
        return ((BonusScoringTier(min_count=1, max_count=None, points=pts),), True)

    # Tiered: "X to Y birds: P1; Z+ birds: P2" (possibly with /swift_start variant)
    # Also handles "3 consecutive birds: 3; 4 consecutive birds: 5; 5 consecutive birds: 8"
    # And "One set: 3; Two sets: 8"
    # And "2 to 3 types: 3; 4 types: 5; 5 types: 8"
    tiers = []
    segments = text.split(";")
    for seg in segments:
        seg = seg.strip()

        # "X to Y <unit>: P" or "X-Y <unit>: P"
        m = re.match(r"(\d+)\s+to\s+(\d+)\s+\w+:\s*(\d+)(?:/\d+)?", seg)
        if m:
            tiers.append(BonusScoringTier(
                min_count=int(m.group(1)),
                max_count=int(m.group(2)),
                points=int(m.group(3)),
            ))
            continue

        # "Z+ <unit>: P"
        m = re.match(r"(\d+)\+\s+\w+:\s*(\d+)(?:/\d+)?", seg)
        if m:
            tiers.append(BonusScoringTier(
                min_count=int(m.group(1)),
                max_count=None,
                points=int(m.group(2)),
            ))
            continue

        # "N <unit>: P" (exact count, like "5 birds: 8" or "5 consecutive birds: 8")
        m = re.match(r"(\d+)\s+\w+(?:\s+\w+)?:\s*(\d+)(?:/\d+)?", seg)
        if m:
            tiers.append(BonusScoringTier(
                min_count=int(m.group(1)),
                max_count=int(m.group(1)),
                points=int(m.group(2)),
            ))
            continue

        # "One set: P" / "Two sets: P"
        m = re.match(r"(One|Two|Three)\s+\w+:\s*(\d+)", seg, re.IGNORECASE)
        if m:
            word_to_num = {"one": 1, "two": 2, "three": 3}
            tiers.append(BonusScoringTier(
                min_count=word_to_num[m.group(1).lower()],
                max_count=word_to_num[m.group(1).lower()],
                points=int(m.group(2)),
            ))
            continue

    if tiers:
        return (tuple(tiers), False)

    return ((), False)


def _parse_bonus_sets(set_text: str) -> frozenset[GameSet]:
    """Parse bonus card set(s) from text like 'core', 'core, asia', 'americas'."""
    sets = set()
    for part in set_text.split(","):
        part = part.strip().lower()
        if part in SET_NAME_MAP:
            sets.add(SET_NAME_MAP[part])
    return frozenset(sets)


def load_bonus_cards(filepath: Path) -> list[BonusCard]:
    """Load bonus cards from the Excel file, filtered to included sets."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["Bonus"]

    cards = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # Skip header

        if len(row) < 5:
            continue

        name = row[0]
        if not name:
            continue

        name_str = str(name).strip()

        # Skip automa-only cards
        if name_str.startswith("[automa]"):
            continue

        set_text = str(row[1] or "").strip()
        game_sets = _parse_bonus_sets(set_text)

        # Skip if none of the card's sets are in our included sets
        included_game_sets = {GameSet.CORE, GameSet.EUROPEAN, GameSet.OCEANIA, GameSet.ASIA}
        if not game_sets & included_game_sets:
            continue

        is_automa = bool(row[2] and str(row[2]).strip())
        vp_text = str(row[4] or "").strip()
        scoring_tiers, is_per_bird = _parse_bonus_vp(vp_text)

        # Clean up name (remove [swift_start_asia] tags)
        clean_name = re.sub(r"\s*\[.*?\]", "", name_str).strip()

        pct = row[6] if len(row) > 6 else None
        draft_pct = float(pct) / 100 if pct and isinstance(pct, (int, float)) else None

        card = BonusCard(
            name=clean_name,
            game_sets=game_sets,
            condition_text=str(row[3] or "").strip(),
            explanation_text=str(row[5] or "").strip() if row[5] else None,
            scoring_tiers=scoring_tiers,
            is_per_bird=is_per_bird,
            is_automa=is_automa,
            draft_value_pct=draft_pct,
        )
        cards.append(card)

    wb.close()
    return cards


def load_goals(filepath: Path) -> list[Goal]:
    """Load end-of-round goals from the Excel file, filtered to included sets.

    Excludes Asia duet goals (which have None scoring values).
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["Goals"]

    goals = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # Skip header

        if len(row) < 6:
            continue

        description = row[0]
        if not description:
            continue

        set_text = str(row[1] or "").strip().lower()
        if set_text not in SET_NAME_MAP:
            continue

        game_set = SET_NAME_MAP[set_text]

        # Skip if not in included sets
        if game_set not in {GameSet.CORE, GameSet.EUROPEAN, GameSet.OCEANIA, GameSet.ASIA}:
            continue

        # Parse scoring (columns 2-5 = 4th, 3rd, 2nd, 1st place)
        scoring_vals = row[2:6]

        # Skip Asia duet goals (all None scoring)
        if all(v is None for v in scoring_vals):
            continue

        # Handle "no goal" entry (Oceania — awards nectar instead)
        if str(description).strip() == "no goal":
            continue

        scoring = tuple(
            float(v) if v is not None else 0.0
            for v in scoring_vals
        )

        reverse = str(row[6] or "").strip()

        goals.append(Goal(
            description=str(description).strip(),
            game_set=game_set,
            scoring=scoring,
            reverse_description=reverse,
        ))

    wb.close()
    return goals
